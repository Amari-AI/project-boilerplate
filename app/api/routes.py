from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from typing import List, Dict, Any, Optional
import os
import tempfile
import base64
import io
import uuid
from pydantic import BaseModel

from app.services.document_processor import process_documents
from app.services.llm_service import extract_shipment_data
from app.core.config import settings

router = APIRouter()

# Store uploaded documents temporarily (in production, use proper storage)
document_storage = {}

class ShipmentData(BaseModel):
    bill_of_lading_number: Optional[str] = None
    container_number: Optional[str] = None
    consignee_name: Optional[str] = None
    consignee_address: Optional[str] = None
    date: Optional[str] = None
    line_items_count: Optional[int] = None
    average_gross_weight: Optional[str] = None
    average_price: Optional[str] = None


class ProcessDocumentsResponse(BaseModel):
    status: str
    shipment_data: Dict[str, Any]
    metadata: Dict[str, Any]
    document_ids: List[str]

class UpdateShipmentRequest(BaseModel):
    shipment_data: ShipmentData
    session_id: str


@router.post("/process-documents", response_model=ProcessDocumentsResponse)
async def process_documents_endpoint(
    files: List[UploadFile] = File(...)
):
    """
    Process multiple document files (PDFs and Excel files) and extract shipment data.
    
    Args:
        files: List of uploaded files (PDFs and/or Excel files)
        
    Returns:
        ProcessDocumentsResponse: Extracted shipment data and metadata
    """
    
    if not files:
        raise HTTPException(status_code=400, detail="No files provided")
    
    temp_file_paths = []
    
    try:
        # Validate and save uploaded files
        for file in files:
            # Check file extension
            file_ext = os.path.splitext(file.filename)[1].lower()
            if file_ext not in settings.ALLOWED_DOCUMENT_TYPES:
                raise HTTPException(
                    status_code=400, 
                    detail=f"File type {file_ext} not supported. Allowed types: {settings.ALLOWED_DOCUMENT_TYPES}"
                )
            
            # Save uploaded file temporarily
            with tempfile.NamedTemporaryFile(suffix=file.filename, delete=False) as temp_file:
                temp_file_paths.append(temp_file.name)
                
                # Write content to temp file
                content = await file.read()
                temp_file.write(content)
        
        # Process documents to extract raw data
        document_data = process_documents(temp_file_paths)
        
        # Use LLM to extract specific shipment fields
        extracted_shipment_data = extract_shipment_data(document_data)
        
        # Store documents for preview
        document_ids = []
        session_id = str(uuid.uuid4())
        
        for i, (path, file) in enumerate(zip(temp_file_paths, files)):
            doc_id = f"{session_id}_{i}"
            with open(path, 'rb') as f:
                file_ext = os.path.splitext(file.filename)[1].lower()
                content_type = 'application/pdf' if file_ext == '.pdf' else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                
                document_storage[doc_id] = {
                    'content': f.read(),
                    'filename': file.filename,  # Use original filename
                    'content_type': content_type
                }
            document_ids.append(doc_id)
        
        # Prepare response
        response = {
            "status": "success",
            "shipment_data": extracted_shipment_data,
            "metadata": {
                "total_files_processed": document_data.get("total_files", 0),
                "pdf_count": document_data.get("pdf_count", 0),
                "excel_count": document_data.get("excel_count", 0),
                "file_types": document_data.get("file_types", []),
                "files": [file.filename for file in files],
                "session_id": session_id
            },
            "document_ids": document_ids
        }
        
        return ProcessDocumentsResponse(**response)
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing documents: {str(e)}")
    
    finally:
        # Clean up temp files
        for path in temp_file_paths:
            try:
                if os.path.exists(path):
                    os.unlink(path)
            except Exception:
                pass


@router.get("/document/{document_id}")
async def get_document(document_id: str):
    """Get a stored document for preview"""
    if document_id not in document_storage:
        raise HTTPException(status_code=404, detail="Document not found")
    
    doc = document_storage[document_id]
    return StreamingResponse(
        io.BytesIO(doc['content']),
        media_type=doc['content_type'],
        headers={"Content-Disposition": f"inline; filename={doc['filename']}"}
    )

@router.get("/document/{document_id}/base64")
async def get_document_base64(document_id: str):
    """Get a stored document as base64 for preview"""
    if document_id not in document_storage:
        raise HTTPException(status_code=404, detail="Document not found")
    
    doc = document_storage[document_id]
    base64_content = base64.b64encode(doc['content']).decode('utf-8')
    
    result = {
        "content": base64_content,
        "content_type": doc['content_type'],
        "filename": doc['filename']
    }
    
    # If it's an Excel file, also provide parsed data for preview
    if doc['content_type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        try:
            import openpyxl
            from io import BytesIO
            
            wb = openpyxl.load_workbook(BytesIO(doc['content']))
            excel_data = {}
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    # Convert row to list and handle None values
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    if any(row_data):  # Only include non-empty rows
                        rows.append(row_data)
                excel_data[sheet_name] = rows[:100]  # Limit to first 100 rows for preview
            
            result['excel_preview'] = excel_data
            wb.close()
        except Exception as e:
            print(f"Error parsing Excel for preview: {e}")
    
    return result

@router.post("/update-shipment")
async def update_shipment_data(request: UpdateShipmentRequest):
    """Update shipment data after manual edits"""
    # In a real application, this would save to a database
    return {
        "status": "success",
        "message": "Shipment data updated successfully",
        "data": request.shipment_data
    }

@router.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "service": "Document Processing API"}