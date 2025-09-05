import openpyxl
import os
from typing import Dict, List

def extract_data_from_excel(file_path: str) -> Dict[str, List[List]]:
    """
    Extract data from an Excel file.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        dict: Dictionary with sheet names as keys and cell data as values
    """
    data = {}
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = []
            
            for row in sheet.iter_rows(values_only=True):
                # Convert None values to empty strings and filter out empty rows
                row_data = [str(cell) if cell is not None else "" for cell in row]
                if any(cell.strip() for cell in row_data):  # Only add non-empty rows
                    sheet_data.append(row_data)
            
            data[sheet_name] = sheet_data
            
        workbook.close()
    except Exception as e:
        print(f"Error reading Excel file {file_path}: {e}")
        return {}
    
    return data

def format_excel_data_for_llm(data: Dict[str, List[List]]) -> str:
    """
    Format Excel data into a readable string for LLM processing.
    
    Args:
        data: Dictionary with sheet data
        
    Returns:
        str: Formatted string representation of the Excel data
    """
    formatted_text = ""
    
    for sheet_name, rows in data.items():
        formatted_text += f"Sheet: {sheet_name}\n"
        formatted_text += "-" * (len(sheet_name) + 7) + "\n"
        
        for i, row in enumerate(rows):
            if i == 0:  # Header row
                formatted_text += " | ".join(row) + "\n"
                formatted_text += "-" * (len(" | ".join(row))) + "\n"
            else:
                formatted_text += " | ".join(row) + "\n"
        
        formatted_text += "\n"
    
    return formatted_text