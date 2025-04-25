# import PyPDF2
import os
from typing import Optional, List
from openpyxl import load_workbook

import pdfplumber
from pdf2image import convert_from_path
import pytesseract

def extract_text_from_pdf(file_path: str, min_text_threshold: int = 100) -> str:
    """
    Extracts text from a PDF file. Uses direct text extraction first; if insufficient,
    falls back to OCR.

    Args:
        file_path: Path to the PDF file.
        min_text_threshold: Minimum number of characters to accept text-based extraction.

    Returns:
        str: Extracted text from the PDF.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    text = ""
    
    # Attempt text-based extraction
    try:
        with pdfplumber.open(file_path) as pdf:
            print(f"PDF has {len(pdf.pages)} pages")
            for i, page in enumerate(pdf.pages):
                page_text = page.extract_text() or ""
                print(f"Text Extraction - Page {i+1} has {len(page_text)} characters")
                text += page_text
    except Exception as e:
        print(f"Text extraction error: {e}")

    # Fall back to OCR if text is too short
    if len(text.strip()) < min_text_threshold:
        print("Text too short, falling back to OCR...")
        text = ""
        try:
            images = convert_from_path(file_path)
            for i, img in enumerate(images):
                page_text = pytesseract.image_to_string(img)
                print(f"OCR - Page {i+1} has {len(page_text)} characters")
                text += page_text
        except Exception as e:
            print(f"OCR extraction error: {e}")

    return text

def extract_data_from_excel(file_path: str):
    """
    Extract data from all sheets in an Excel file using openpyxl.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        dict: Dictionary with sheet names as keys and lists of rows as values
    """
    try:
        # Load the workbook
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        
        # Dictionary to store data from all sheets
        all_sheets_data = {}
        
        # Iterate through all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Extract data from current sheet
            sheet_data = []
            for row in ws.rows:
                sheet_data.append([cell.value for cell in row])
            
            # Store data in dictionary
            all_sheets_data[sheet_name] = sheet_data
            print(f"Sheet '{sheet_name}': {len(sheet_data)} rows")
        
        print(f"Successfully extracted data from {len(all_sheets_data)} sheets")
        return all_sheets_data
    
    except Exception as e:
        print(f"Error extracting data from Excel: {e}")
        return None