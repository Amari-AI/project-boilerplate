import openpyxl
from typing import Dict, Any, List

def extract_text_from_xlsx(file_path: str) -> Dict[str, Any]:
    """
    Extract data from an XLSX file.
    
    Args:
        file_path: Path to the XLSX file
        
    Returns:
        dict: Extracted data from all sheets
    """
    data = {}
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = []
            
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    sheet_data.append([str(cell) if cell is not None else "" for cell in row])
            
            data[sheet_name] = sheet_data
            
        workbook.close()
    except Exception as e:
        print(f"Error extracting data from XLSX {file_path}: {str(e)}")
        return {}
    
    return data

def parse_excel_to_json(file_path: str) -> Dict[str, Any]:
    """
    Parse Excel file to structured JSON format for frontend rendering.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        dict: Structured data with headers and rows for each sheet
    """
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        result = {}
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            all_rows = list(sheet.iter_rows(values_only=True))
            
            if not all_rows:
                result[sheet_name] = {
                    "headers": [],
                    "rows": [],
                    "total_rows": 0,
                    "total_columns": 0
                }
                continue
            
            # Filter out empty rows
            non_empty_rows = [row for row in all_rows if any(cell is not None for cell in row)]
            
            if not non_empty_rows:
                result[sheet_name] = {
                    "headers": [],
                    "rows": [],
                    "total_rows": 0,
                    "total_columns": 0
                }
                continue
            
            # Use first row as headers
            headers = [str(cell) if cell is not None else f"Column {i+1}" for i, cell in enumerate(non_empty_rows[0])]
            
            # Get data rows (skip header row)
            rows = []
            for row in non_empty_rows[1:]:
                row_data = [str(cell) if cell is not None else "" for cell in row]
                # Ensure row has same length as headers
                while len(row_data) < len(headers):
                    row_data.append("")
                rows.append(row_data)
            
            result[sheet_name] = {
                "headers": headers,
                "rows": rows,
                "total_rows": len(rows),
                "total_columns": len(headers)
            }
        
        workbook.close()
        return result
        
    except Exception as e:
        print(f"Error parsing Excel to JSON from {file_path}: {str(e)}")
        return {}