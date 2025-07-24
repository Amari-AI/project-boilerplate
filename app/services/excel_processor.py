import openpyxl


def excel_worksheet_to_csv_string(
    workbook: openpyxl.Workbook, sheet_name: str | None = None
) -> str:
    try:
        worksheet = workbook[sheet_name]
    except KeyError:
        return ""

    csv_rows = []

    for row in worksheet.iter_rows(values_only=True):
        # Handle None values and escape commas/quotes
        csv_cells = []
        for cell in row:
            if cell is None:
                csv_cells.append("")
            else:
                cell_str = str(cell)
                # Escape quotes and wrap in quotes if contains comma or quote
                if "," in cell_str or '"' in cell_str or "\n" in cell_str:
                    cell_str = '"' + cell_str.replace('"', '""') + '"'
                csv_cells.append(cell_str)

        csv_rows.append(",".join(csv_cells))

    return "\n".join(csv_rows)


def excel_to_csv_strings_all_sheets(excel_file_path: str) -> dict[str, str]:
    """Convert all sheets in an Excel file to a list of CSV strings"""
    workbook = openpyxl.load_workbook(excel_file_path)
    csv_strings = {}

    for sheet_name in workbook.sheetnames:
        csv_strings[sheet_name] = excel_worksheet_to_csv_string(workbook, sheet_name)

    return csv_strings
