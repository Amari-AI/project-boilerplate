import sys
import pytest
import tempfile
import os
import json
from pathlib import Path
from fastapi.testclient import TestClient
from unittest.mock import patch, MagicMock

# Add the project root directory to Python path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from app.main import app
from app.services.storage_service import JSONStorageService

@pytest.fixture
def client():
    """Create a test client for the FastAPI application."""
    return TestClient(app)

@pytest.fixture
def temp_storage_file():
    """Create a temporary storage file for testing."""
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
        f.write('{"documents": {}, "metadata": {"created_at": "2023-01-01T00:00:00", "last_updated": "2023-01-01T00:00:00"}}')
        temp_file = f.name
    
    yield temp_file
    
    # Cleanup
    if os.path.exists(temp_file):
        os.unlink(temp_file)

@pytest.fixture
def temp_storage_service(temp_storage_file):
    """Create a temporary storage service for testing."""
    return JSONStorageService(temp_storage_file)

@pytest.fixture
def temp_upload_dir():
    """Create a temporary upload directory for testing."""
    temp_dir = tempfile.mkdtemp()
    yield Path(temp_dir)
    
    # Cleanup
    import shutil
    shutil.rmtree(temp_dir, ignore_errors=True)

@pytest.fixture
def sample_pdf_file():
    """Create a sample PDF file for testing."""
    # Simple PDF content (minimal valid PDF)
    pdf_content = b"""%PDF-1.4
1 0 obj
<<
/Type /Catalog
/Pages 2 0 R
>>
endobj
2 0 obj
<<
/Type /Pages
/Kids [3 0 R]
/Count 1
>>
endobj
3 0 obj
<<
/Type /Page
/Parent 2 0 R
/MediaBox [0 0 612 792]
/Contents 4 0 R
>>
endobj
4 0 obj
<<
/Length 44
>>
stream
BT
/F1 12 Tf
100 700 Td
(Test PDF) Tj
ET
endstream
endobj
xref
0 5
0000000000 65535 f 
0000000010 00000 n 
0000000053 00000 n 
0000000104 00000 n 
0000000178 00000 n 
trailer
<<
/Size 5
/Root 1 0 R
>>
startxref
268
%%EOF"""
    
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as f:
        f.write(pdf_content)
        temp_file = f.name
    
    yield temp_file
    
    if os.path.exists(temp_file):
        os.unlink(temp_file)

@pytest.fixture
def sample_excel_file():
    """Create a sample Excel file for testing."""
    try:
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        
        # Add some test data
        ws['A1'] = 'Name'
        ws['B1'] = 'Value'
        ws['A2'] = 'Test Item'
        ws['B2'] = '123'
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            temp_file = f.name
            
        yield temp_file
        
        if os.path.exists(temp_file):
            os.unlink(temp_file)
            
    except ImportError:
        pytest.skip("openpyxl not available")

@pytest.fixture
def mock_llm_service():
    """Mock the LLM service for testing."""
    mock_data = {
        "bill_of_lading_number": "TEST123",
        "container_number": "CONT456",
        "consignee_name": "Test Company",
        "consignee_address": "123 Test St",
        "date": "2023-01-01",
        "line_items_count": "5",
        "average_gross_weight": "100.0",
        "average_price": "50.0"
    }
    
    with patch('app.services.llm_service.extract_field_from_document', return_value=mock_data):
        yield mock_data 